#!/usr/bin/env python3
"""本地 Excel 数据提取工具 - 从 .xlsx 文件中提取缺陷数据

支持两种图片来源：
1. Excel 内嵌图片（自动提取，按锚定位置匹配行）
2. 外部 images/ 目录（按 r{行号}_{字段}_ 命名规则匹配）

用法：
    python scripts/extract_excel.py --input defects.xlsx --config config.yaml
"""

import os
import sys
import argparse

# 将 scripts/ 加入 path 以导入 utils
sys.path.insert(0, os.path.dirname(__file__))
from utils import load_config, build_defect_data, save_results


def find_column_indices(ws, column_names):
    """根据列名找到列索引（第一行为表头）"""
    print(f"\n=== 查找列索引 ===")
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else "")
    
    indices = {}
    for name, col_name in column_names.items():
        try:
            idx = headers.index(col_name)
            indices[name] = idx
            print(f"  {name}: 列{chr(65+idx)} = '{col_name}'")
        except ValueError:
            print(f"  ⚠️ 列 '{col_name}' 未找到，跳过 {name}")
    
    return indices


def extract_embedded_images(ws, output_dir):
    """从 .xlsx 中提取嵌入图片，按行号+列字段保存

    Returns:
        dict: { row_idx: { field_name: [filenames] } }
    """
    print(f"\n=== 提取嵌入图片 ===")
    if not hasattr(ws, '_images') or not ws._images:
        print("  📭 Excel 中无嵌入图片")
        return {}

    os.makedirs(output_dir, exist_ok=True)
    embedded = {}
    img_count = 0

    for img in ws._images:
        # 获取图片锚定的行号（1-based）
        try:
            anchor = img.anchor
            if hasattr(anchor, '_from'):
                row_0based = anchor._from.row  # 0-based
                col_0based = anchor._from.col  # 0-based
            elif isinstance(anchor, str):
                # 字符串锚点如 "B2"，解析行号
                import re
                m = re.match(r'([A-Z]+)(\d+)', anchor)
                if m:
                    row_0based = int(m.group(2)) - 1
                    col_0based = ord(m.group(1)[0]) - ord('A')
                else:
                    continue
            else:
                continue

            row_idx = row_0based + 1  # 转为 1-based（与 Excel 行号一致）

            # 获取图片数据
            img_data = img._data() if hasattr(img, '_data') and callable(img._data) else None
            if img_data is None and hasattr(img, 'ref'):
                try:
                    img_data = img.ref.getvalue() if hasattr(img.ref, 'getvalue') else img.ref.read()
                except Exception:
                    pass
            if img_data is None:
                continue

            # 保存到 output_dir
            filename = f"r{row_idx}_embedded_col{col_0based}_{img_count}.png"
            filepath = os.path.join(output_dir, filename)
            with open(filepath, 'wb') as f:
                f.write(img_data)

            embedded.setdefault(row_idx, {}).setdefault(f"_col{col_0based}", []).append(filename)
            img_count += 1

        except Exception as e:
            print(f"  ⚠️ 图片提取失败: {e}")
            continue

    print(f"  ✅ 共提取 {img_count} 张嵌入图片")
    return embedded


def map_images_to_fields(embedded, column_indices):
    """将嵌入图片按列号映射到字段（screenshot/design_ref）

    Args:
        embedded: { row_idx: { "_col0": [files] } }
        column_indices: { "screenshot": 2, "design_ref": 5, ... }

    Returns:
        { row_idx: { "screenshot": [files], "design_ref": [files] } }
    """
    # 构建 col_idx → field_name 的反向映射
    col_to_field = {}
    for field, col_idx in column_indices.items():
        if field in ("screenshot", "design_ref"):
            col_to_field[col_idx] = field

    result = {}
    for row_idx, cols in embedded.items():
        for col_key, files in cols.items():
            # col_key 格式 "_col{N}"
            try:
                col_idx = int(col_key.replace("_col", ""))
            except ValueError:
                continue

            field = col_to_field.get(col_idx)
            if not field:
                # 没有精确匹配，默认归入 screenshot
                field = "screenshot"

            result.setdefault(row_idx, {}).setdefault(field, []).extend(files)

    return result


def read_excel_data(ws, column_indices, images_dir=None, embedded_map=None):
    """读取 Excel 数据"""
    print(f"\n=== 读取 Excel 数据 ===")
    defects = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
        
        # 跳过空行
        if not any(values):
            continue
        
        defect = {"row": row_idx}
        has_data = False
        
        for field, col_idx in column_indices.items():
            if col_idx < len(values):
                defect[field] = values[col_idx]
                if defect[field]:
                    has_data = True
            else:
                defect[field] = ""
        
        # 合并图片来源：嵌入图片 + 外部目录
        all_images = {}

        # 1) 嵌入图片
        if embedded_map and row_idx in embedded_map:
            for field, files in embedded_map[row_idx].items():
                all_images.setdefault(field, []).extend(
                    [os.path.join(images_dir or "images", f) for f in files]
                )

        # 2) 外部目录（按命名规则）
        if images_dir and os.path.isdir(images_dir):
            for field in ["screenshot", "design_ref"]:
                pattern = f"r{row_idx}_{field}_"
                matching = [f for f in os.listdir(images_dir) if f.startswith(pattern)]
                if matching:
                    all_images.setdefault(field, []).extend(
                        [os.path.join(images_dir, f) for f in matching]
                    )

        # 写入 defect
        for field, files in all_images.items():
            key = f"{field}_files"
            defect[key] = files
            print(f"  📎 row{row_idx} {field}: {len(files)} 张图片")

        if has_data and defect.get("description"):
            defects.append(defect)
    
    print(f"  ✅ 共提取 {len(defects)} 条缺陷数据")
    return defects


def main():
    parser = argparse.ArgumentParser(description="从本地 Excel 提取缺陷数据")
    parser.add_argument("--input", required=True, help="Excel 文件路径 (.xlsx)")
    parser.add_argument("--config", default="config.yaml", help="配置文件路径")
    parser.add_argument("--output", default="pending_defects.json", help="输出文件路径")
    parser.add_argument("--images-dir", default="images", help="图片输出目录（嵌入图片提取至此，外部图片也从此目录匹配）")
    parser.add_argument("--sheet", default=0, type=int, help="工作表索引（默认第一个）")
    args = parser.parse_args()
    
    # 检查文件
    if not os.path.isfile(args.input):
        print(f"❌ 文件不存在: {args.input}")
        return
    
    # 检查 openpyxl
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ 缺少 openpyxl，请运行: pip install openpyxl")
        return
    
    config = load_config(args.config)
    
    print(f"\n=== 打开 Excel: {args.input} ===")
    wb = load_workbook(args.input)  # 不传 data_only=True，以获取嵌入图片
    
    if args.sheet >= len(wb.sheetnames):
        print(f"❌ 工作表索引 {args.sheet} 超出范围，共 {len(wb.sheetnames)} 个")
        return
    
    ws = wb[wb.sheetnames[args.sheet]]
    print(f"  📋 工作表: {wb.sheetnames[args.sheet]}")
    
    column_names = config.get("column_mapping", config.get("feishu_columns", {}))
    column_indices = find_column_indices(ws, column_names)
    
    # 提取嵌入图片
    embedded = extract_embedded_images(ws, args.images_dir)
    embedded_map = map_images_to_fields(embedded, column_indices)

    # 读取数据（含图片关联）
    defects = read_excel_data(ws, column_indices, args.images_dir, embedded_map)
    results = build_defect_data(defects, config)
    
    save_results(results, args.output)
    
    print(f"\n=== 完成 ===")
    total_imgs = sum(
        len(d.get("screenshot_files", [])) + len(d.get("design_ref_files", []))
        for d in results
    )
    print(f"✅ 共 {len(results)} 条缺陷，{total_imgs} 张图片")


if __name__ == "__main__":
    main()
